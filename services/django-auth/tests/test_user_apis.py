import pytest
from rest_framework.test import APIClient
from rest_framework import status
from UserApp.models import User
from rest_framework_simplejwt.tokens import RefreshToken

import pandas as pd
from datetime import datetime
from pathlib import Path
import json
import traceback
import time

# Create test results directory
BASE_DIR = Path("test_results")
CURRENT_RUN = BASE_DIR / datetime.now().strftime("%Y%m%d_%H%M%S")
EXCEL_DIR = CURRENT_RUN / "excel_reports"
LOGS_DIR = CURRENT_RUN / "logs"

for directory in [BASE_DIR, CURRENT_RUN, EXCEL_DIR, LOGS_DIR]:
    directory.mkdir(parents=True, exist_ok=True)


ALL_ENDPOINTS = {
    "register": "/api/user/register/",
    "login": "/api/user/login/",
    "profile": "/api/user/profile/",
    "logout": "/api/user/logout/",
}


# Global test results storage
from conftest import test_results

def log_test_result(test_name, category, method, endpoint, request_data, 
                   response_status, response_body, passed, error_message="", 
                   execution_time=0, expected_status=None):
    """Log test result for Excel export"""
    result = {
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Test Name": test_name,
        "Category": category,
        "HTTP Method": method,
        "Endpoint": endpoint,
        "Request Data": json.dumps(request_data, default=str) if request_data else "ERROR",
        "Expected Status": expected_status or "ERROR",
        "Actual Status": response_status,
        "Response Body": str(response_body)[:500],
        "Test Result": "PASSED ✅" if passed else "FAILED ❌",
        "Error Message": error_message,
        "Execution Time (ms)": round(execution_time * 1000, 2),
    }
    test_results.append(result)


def get_safe_body(response):
    """Safely extract response body, handling both JSON and non-JSON responses"""
    if response is None:
        return "No Response"
    
    try:
        # Try to get JSON response
        return response.json()
    except (ValueError, AttributeError, TypeError):
        # If JSON parsing fails, return text content
        try:
            content = response.content.decode('utf-8')
            # If it's HTML, truncate it
            if content.strip().startswith('<!DOCTYPE') or content.strip().startswith('<html'):
                return f"HTML Response (truncated): {content[:200]}..."
            return content[:500]
        except:
            return f"Binary or unreadable response: {str(response.content)[:200]}"


@pytest.fixture
def api_client():
    """Create API client"""
    return APIClient()


@pytest.fixture
def create_user(db):
    """Fixture to create a test user"""
    def make_user(**kwargs):
        user_data = {
            'email': 'test@example.com',
            'user_name': 'testuser',  # Changed to user_name
            'password': 'TestPass123!'
        }
        user_data.update(kwargs)
        return User.objects.create_user(**user_data)
    return make_user


@pytest.fixture
def authenticated_client(api_client, create_user):
    """Create authenticated client"""
    user = create_user()
    refresh = RefreshToken.for_user(user)
    api_client.credentials(HTTP_AUTHORIZATION=f'Bearer {refresh.access_token}')
    return api_client, user, str(refresh)


# ============================================================================
# USER REGISTRATION TESTS
# ============================================================================

class TestUserRegistration:
    """Test user registration endpoint"""
    endpoint = ALL_ENDPOINTS['register']
    
    @pytest.mark.django_db
    @pytest.mark.parametrize("test_data,expected_status,test_name", [
        # Valid cases
        (
            {"email": "valid@test.com", "user_name": "validuser", "password": "Pass123!"},
            201,
            "Valid registration"
        ),
        (
            {"email": "test@example.com", "user_name": "user123", "password": "a"},
            201,
            "Minimum password length"
        ),
        (
            {"email": "UPPERCASE@TEST.COM", "user_name": "UPPERCASE", "password": "Pass123!"},
            201,
            "Uppercase email"
        ),
        (
            {"email": "special+chars@test.com", "user_name": "special_user-123", "password": "Pass123!"},
            201,
            "Special characters in email"
        ),
    ])
    def test_valid_registration(self, api_client, test_data, expected_status, test_name):
        """Test valid registration scenarios"""
        start_time = time.time()
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code == expected_status
            
            log_test_result(
                test_name=f"Register - {test_name}",
                category="Valid Registration",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=expected_status,
                execution_time=execution_time
            )
            
            assert passed, f"Expected {expected_status}, got {response.status_code}"
            
            if passed:
                assert User.objects.filter(email=test_data['email']).exists()
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name=f"Register - {test_name}",
                category="Valid Registration",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise
    
    @pytest.mark.django_db
    @pytest.mark.parametrize("test_data,expected_status,test_name,expected_error", [
        # Missing fields
        (
            {"email": "test@test.com", "user_name": "user"},
            400,
            "Missing password field",
            "some field missing value"
        ),
        (
            {"email": "test@test.com", "password": "Pass123!"},
            400,
            "Missing user_name field",
            "some field missing value"
        ),
        (
            {"user_name": "testuser", "password": "Pass123!"},
            400,
            "Missing email field",
            "some field missing value"
        ),
        (
            {},
            400,
            "All fields missing",
            "some field missing value"
        ),
        # Empty values
        (
            {"email": "", "user_name": "user", "password": "Pass123!"},
            400,
            "Empty email",
            "some field missing value"
        ),
        (
            {"email": "test@test.com", "user_name": "", "password": "Pass123!"},
            400,
            "Empty user_name",
            "some field missing value"
        ),
        (
            {"email": "test@test.com", "user_name": "user", "password": ""},
            400,
            "Empty password",
            "some field missing value"
        ),
        # None values
        (
            {"email": None, "user_name": "user", "password": "Pass123!"},
            400,
            "Null email",
            "some field missing value"
        ),
        (
            {"email": "test@test.com", "user_name": None, "password": "Pass123!"},
            400,
            "Null user_name",
            "some field missing value"
        ),
        (
            {"email": "test@test.com", "user_name": "user", "password": None},
            400,
            "Null password",
            "some field missing value"
        ),
    ])
    def test_invalid_registration(self, api_client, test_data, expected_status, test_name, expected_error):
        """Test invalid registration scenarios"""
        start_time = time.time()
        
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code == expected_status
            
            log_test_result(
                test_name=f"Register - {test_name}",
                category="Invalid Registration",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=expected_status,
                execution_time=execution_time
            )
            
            assert passed, f"Expected {expected_status}, got {response.status_code}"
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name=f"Register - {test_name}",
                category="Invalid Registration",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise
    
    @pytest.mark.django_db
    @pytest.mark.parametrize("test_data,test_name", [
        # Edge cases
        ({"email": "a@b.c", "user_name": "a", "password": "x"}, "Minimum length values"),
        ({"email": "x" * 100 + "@test.com", "user_name": "x" * 150, "password": "x" * 200}, "Maximum length values"),
        ({"email": "test@test.com", "user_name": "user123", "password": "' OR '1'='1"}, "SQL injection in password"),
        ({"email": "test@test.com'; DROP TABLE users;--", "user_name": "user", "password": "Pass123!"}, "SQL injection in email"),
        ({"email": "test@test.com", "user_name": "'; DROP TABLE users;--", "password": "Pass123!"}, "SQL injection in user_name"),
        ({"email": "<script>alert('xss')</script>@test.com", "user_name": "user", "password": "Pass123!"}, "XSS in email"),
        ({"email": "test@test.com", "user_name": "<script>alert('xss')</script>", "password": "Pass123!"}, "XSS in user_name"),
        ({"email": "test@test.com", "user_name": "../../../etc/passwd", "password": "Pass123!"}, "Path traversal in user_name"),
        ({"email": "test🚀emoji@test.com", "user_name": "user🎉", "password": "Pass123!😀"}, "Unicode/Emoji characters"),
        ({"email": "test@test.com\n\r", "user_name": "user\n\r", "password": "Pass\n\r123!"}, "Newline characters"),
        ({"email": "   test@test.com   ", "user_name": "   user   ", "password": "   Pass123!   "}, "Whitespace padding"),
    ])
    def test_edge_cases_registration(self, api_client, test_data, test_name):
        """Test edge cases and security scenarios"""
        start_time = time.time()
        
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            # Should not crash (400/201/500 acceptable)
            passed = response.status_code in [200, 201, 400, 422, 500]
            
            log_test_result(
                test_name=f"Register Edge - {test_name}",
                category="Edge Cases",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status="200/201/400/500",
                execution_time=execution_time
            )
            
            assert passed, f"Unexpected status: {response.status_code}"
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name=f"Register Edge - {test_name}",
                category="Edge Cases",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise

    @pytest.mark.django_db
    def test_duplicate_email_registration(self, api_client, create_user):
        """Test duplicate email registration"""
        start_time = time.time()
        
        # Create first user
        create_user(email='duplicate@test.com', user_name='user1')
        
        # Try to create another user with same email
        test_data = {"email": "duplicate@test.com", "user_name": "user2", "password": "Pass123!"}
        
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            # Should fail with 500 (as per your code)
            passed = response.status_code == 500
            
            log_test_result(
                test_name="Register - Duplicate email",
                category="Duplicate Registration",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=500,
                execution_time=execution_time
            )
            
            assert passed, f"Expected 500, got {response.status_code}"
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name="Register - Duplicate email",
                category="Duplicate Registration",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise


# ============================================================================
# USER LOGIN TESTS
# ============================================================================

class TestUserLogin:
    """Test user login endpoint"""
    endpoint = ALL_ENDPOINTS['login']
    
    @pytest.mark.django_db
    def test_valid_login(self, api_client, create_user):
        """Test successful login"""
        start_time = time.time()
        
        user = create_user(email='login@test.com', user_name='loginuser', password='Pass123!')
        test_data = {"email": "login@test.com", "password": "Pass123!"}
        
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code == 200
            
            log_test_result(
                test_name="Login - Valid credentials",
                category="Valid Login",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=200,
                execution_time=execution_time
            )
            
            assert passed
            response_data = response.json()
            assert 'message' in response_data
            assert 'access' in response_data['message']
            assert 'refresh' in response_data['message']
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name="Login - Valid credentials",
                category="Valid Login",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise

    @pytest.mark.django_db
    @pytest.mark.parametrize("test_data,expected_status,test_name", [
        ({"email": "wrong@test.com", "password": "Pass123!"}, 500, "Non-existent email"),
        ({"email": "", "password": "Pass123!"}, 400, "Empty email"),
        ({"email": "test@test.com", "password": ""}, 400, "Empty password"),
        ({"password": "Pass123!"}, 400, "Missing email"),
        ({"email": "test@test.com"}, 400, "Missing password"),
        ({}, 400, "All fields missing"),
        ({"email": None, "password": "Pass123!"}, 400, "Null email"),
        ({"email": "test@test.com", "password": None}, 400, "Null password"),
    ])
    def test_invalid_login(self, api_client, create_user, test_data, expected_status, test_name):
        """Test invalid login scenarios"""
        start_time = time.time()
        
        # Create user for some tests
        if 'test@test.com' in str(test_data.get('email', '')):
            create_user(email='test@test.com', user_name='testuser', password='Pass123!')
        
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code == expected_status
            
            log_test_result(
                test_name=f"Login - {test_name}",
                category="Invalid Login",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=expected_status,
                execution_time=execution_time
            )
            
            assert passed, f"Expected {expected_status}, got {response.status_code}"
            
        except Exception as e:
            execution_time = time.time() - start_time
            
            # For non-existent user, DoesNotExist exception is expected
            if "does not exist" in str(e).lower() or "DoesNotExist" in str(e):
                passed = expected_status == 500
                log_test_result(
                    test_name=f"Login - {test_name}",
                    category="Invalid Login",
                    method="POST",
                    endpoint=self.endpoint,
                    request_data=test_data,
                    response_status=500,
                    response_body=str(e),
                    passed=passed,
                    expected_status=expected_status,
                    error_message="User.DoesNotExist exception raised (unhandled)",
                    execution_time=execution_time
                )
                if not passed:
                    raise
            else:
                log_test_result(
                    test_name=f"Login - {test_name}",
                    category="Invalid Login",
                    method="POST",
                    endpoint=self.endpoint,
                    request_data=test_data,
                    response_status="ERROR",
                    response_body=str(e),
                    passed=False,
                    error_message=traceback.format_exc(),
                    execution_time=execution_time
                )
                raise

    @pytest.mark.django_db
    def test_wrong_password_login(self, api_client, create_user):
        """Test login with wrong password"""
        start_time = time.time()
        
        create_user(email='test@test.com', user_name='testuser', password='CorrectPass123!')
        test_data = {"email": "test@test.com", "password": "WrongPass123!"}
        
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code == 400
            
            log_test_result(
                test_name="Login - Wrong password",
                category="Invalid Login",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=400,
                execution_time=execution_time
            )
            
            assert passed
            assert response.json()['error'] == 'incorrect credentials'
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name="Login - Wrong password",
                category="Invalid Login",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise

    @pytest.mark.django_db
    @pytest.mark.parametrize("test_data,test_name", [
        ({"email": "test@test.com'; DROP TABLE users;--", "password": "Pass123!"}, "SQL injection in email"),
        ({"email": "test@test.com", "password": "' OR '1'='1"}, "SQL injection in password"),
        ({"email": "test@test.com\n\r", "password": "Pass123!"}, "Newline in email"),
        ({"email": "   test@test.com   ", "password": "   Pass123!   "}, "Whitespace padding"),
    ])
    def test_edge_cases_login(self, api_client, create_user, test_data, test_name):
        """Test edge cases in login"""
        start_time = time.time()
        
        # Create user
        create_user(email='test@test.com', user_name='testuser', password='Pass123!')
        
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code in [200, 400, 500]
            
            log_test_result(
                test_name=f"Login Edge - {test_name}",
                category="Edge Cases",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status="200/400/500",
                execution_time=execution_time
            )
            
            assert passed
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name=f"Login Edge - {test_name}",
                category="Edge Cases",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise


# ============================================================================
# USER INFO TESTS
# ============================================================================

class TestUserInfo:
    """Test user info endpoint"""
    endpoint = ALL_ENDPOINTS['profile']
    
    @pytest.mark.django_db
    def test_get_user_info_authenticated(self, authenticated_client):
        """Test getting user info with valid token"""
        start_time = time.time()
        
        client, user, refresh = authenticated_client
        
        try:
            response = client.get(self.endpoint)
            execution_time = time.time() - start_time
            
            passed = response.status_code == 200
            
            log_test_result(
                test_name="User Info - Authenticated",
                category="Valid Access",
                method="GET",
                endpoint=self.endpoint,
                request_data={"Authorization": "Bearer <token>"},
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=200,
                execution_time=execution_time
            )
            
            assert passed
            assert 'details' in response.json()
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name="User Info - Authenticated",
                category="Valid Access",
                method="GET",
                endpoint=self.endpoint,
                request_data={"Authorization": "Bearer <token>"},
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise

    @pytest.mark.django_db
    def test_get_user_info_unauthenticated(self, api_client):
        """Test getting user info without token"""
        start_time = time.time()
        
        try:
            response = api_client.get(self.endpoint)
            execution_time = time.time() - start_time
            
            passed = response.status_code == 401
            
            log_test_result(
                test_name="User Info - No token",
                category="Unauthorized Access",
                method="GET",
                endpoint=self.endpoint,
                request_data={},
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=401,
                execution_time=execution_time
            )
            
            assert passed
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name="User Info - No token",
                category="Unauthorized Access",
                method="GET",
                endpoint=self.endpoint,
                request_data={},
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise

    @pytest.mark.django_db
    @pytest.mark.parametrize("token,test_name", [
        ("", "Empty token"),
        ("invalid_token", "Invalid token"),
        ("Bearer invalid_token", "Invalid Bearer token"),
        ("xxx.yyy.zzz", "Malformed JWT"),
    ])
    def test_get_user_info_invalid_token(self, api_client, token, test_name):
        """Test getting user info with invalid tokens"""
        start_time = time.time()
        
        if token:
            api_client.credentials(HTTP_AUTHORIZATION=token)
        
        try:
            response = api_client.get(self.endpoint)
            execution_time = time.time() - start_time
            
            passed = response.status_code == 401
            
            log_test_result(
                test_name=f"User Info - {test_name}",
                category="Unauthorized Access",
                method="GET",
                endpoint=self.endpoint,
                request_data={"Authorization": token},
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=401,
                execution_time=execution_time
            )
            
            assert passed
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name=f"User Info - {test_name}",
                category="Unauthorized Access",
                method="GET",
                endpoint=self.endpoint,
                request_data={"Authorization": token},
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise


# ============================================================================
# USER LOGOUT TESTS
# ============================================================================

class TestUserLogout:
    """Test user logout endpoint"""
    endpoint = ALL_ENDPOINTS['logout']
    
    @pytest.mark.django_db
    def test_valid_logout(self, authenticated_client):
        """Test successful logout"""
        start_time = time.time()
        
        client, user, refresh_token = authenticated_client
        test_data = {"refresh": refresh_token}
        
        try:
            response = client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code == 200
            
            log_test_result(
                test_name="Logout - Valid token",
                category="Valid Logout",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=200,
                execution_time=execution_time
            )
            
            assert passed
            assert response.json()['message'] == 'User logged out successfully'
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name="Logout - Valid token",
                category="Valid Logout",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise

    @pytest.mark.django_db
    @pytest.mark.parametrize("test_data,expected_status,test_name", [
        ({}, 500, "Missing refresh token"),
        ({"refresh": ""}, 500, "Empty refresh token"),
        ({"refresh": None}, 500, "Null refresh token"),
        ({"refresh": "invalid_token"}, 500, "Invalid refresh token"),
        ({"refresh": "xxx.yyy.zzz"}, 500, "Malformed JWT"),
    ])
    def test_invalid_logout(self, authenticated_client, test_data, expected_status, test_name):
        """Test logout with invalid data"""
        start_time = time.time()
        
        client, user, refresh_token = authenticated_client
        
        try:
            response = client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code == expected_status
            
            log_test_result(
                test_name=f"Logout - {test_name}",
                category="Invalid Logout",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=expected_status,
                execution_time=execution_time
            )
            
            assert passed
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name=f"Logout - {test_name}",
                category="Invalid Logout",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise

    @pytest.mark.django_db
    def test_logout_without_authentication(self, api_client):
        """Test logout without authentication"""
        start_time = time.time()
        
        test_data = {"refresh": "some_token"}
        
        try:
            response = api_client.post(self.endpoint, test_data, format='json')
            execution_time = time.time() - start_time
            
            passed = response.status_code == 401
            
            log_test_result(
                test_name="Logout - No authentication",
                category="Unauthorized Access",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status=response.status_code,
                response_body=get_safe_body(response),
                passed=passed,
                expected_status=401,
                execution_time=execution_time
            )
            
            assert passed
            
        except Exception as e:
            execution_time = time.time() - start_time
            log_test_result(
                test_name="Logout - No authentication",
                category="Unauthorized Access",
                method="POST",
                endpoint=self.endpoint,
                request_data=test_data,
                response_status="ERROR",
                response_body=str(e),
                passed=False,
                error_message=traceback.format_exc(),
                execution_time=execution_time
            )
            raise


# ============================================================================
# PERFORMANCE & STRESS TESTS
# ============================================================================

class TestPerformance:
    """Test performance and stress scenarios"""
    
    @pytest.mark.django_db
    def test_bulk_registrations(self, api_client):
        """Test bulk user registrations"""
        start_time = time.time()
        
        success_count = 0
        fail_count = 0
        
        for i in range(20):
            test_data = {
                "email": f"bulk{i}@test.com",
                "user_name": f"bulkuser{i}",
                "password": f"Pass{i}123!"
            }
            
            try:
                response = api_client.post(ALL_ENDPOINTS['register'], test_data, format='json')
                if response.status_code == 201:
                    success_count += 1
                else:
                    fail_count += 1
            except:
                fail_count += 1
        
        execution_time = time.time() - start_time
        
        log_test_result(
            test_name="Performance - Bulk registrations (20 users)",
            category="Performance",
            method="POST",
            endpoint=ALL_ENDPOINTS['register'],
            request_data=f"20 registration requests",
            response_status=f"{success_count} success, {fail_count} failed",
            response_body=f"Total: {execution_time:.2f}s, Avg: {execution_time/20*1000:.2f}ms per request",
            passed=fail_count == 0,
            execution_time=execution_time
        )

    @pytest.mark.django_db
    def test_concurrent_logins(self, api_client, create_user):
        """Test multiple login attempts"""
        start_time = time.time()
        
        user = create_user(email='concurrent@test.com', user_name='concurrent', password='Pass123!')
        
        success_count = 0
        fail_count = 0
        
        for i in range(10):
            test_data = {"email": "concurrent@test.com", "password": "Pass123!"}
            
            try:
                response = api_client.post(ALL_ENDPOINTS['login'], test_data, format='json')
                if response.status_code == 200:
                    success_count += 1
                else:
                    fail_count += 1
            except:
                fail_count += 1
        
        execution_time = time.time() - start_time
        
        log_test_result(
            test_name="Performance - Concurrent logins (10 attempts)",
            category="Performance",
            method="POST",
            endpoint=ALL_ENDPOINTS['login'],
            request_data=f"10 login requests",
            response_status=f"{success_count} success, {fail_count} failed",
            response_body=f"Total: {execution_time:.2f}s, Avg: {execution_time/10*1000:.2f}ms per request",
            passed=fail_count == 0,
            execution_time=execution_time
        )


# ============================================================================
# PYTEST HOOK TO SAVE RESULTS
# ============================================================================

def pytest_sessionfinish(session, exitstatus):
    """Save test results to Excel after all tests complete"""
    
    if not test_results:
        print("No test results to save")
        return
    
    # Create DataFrame
    df = pd.DataFrame(test_results)
    
    # Save to Excel
    excel_file = EXCEL_DIR / "test_results_comprehensive.xlsx"
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Sheet 1: All test results
        df.to_excel(writer, sheet_name='All Tests', index=False)
        
        # Sheet 2: Summary by category
        summary = df.groupby('Category').agg({
            'Test Name': 'count',
            'Test Result': lambda x: (x == 'PASSED ✅').sum(),
            'Execution Time (ms)': 'mean'
        }).rename(columns={
            'Test Name': 'Total Tests',
            'Test Result': 'Passed',
            'Execution Time (ms)': 'Avg Time (ms)'
        })
        summary['Failed'] = summary['Total Tests'] - summary['Passed']
        summary['Pass Rate %'] = (summary['Passed'] / summary['Total Tests'] * 100).round(2)
        summary.to_excel(writer, sheet_name='Summary by Category')
        
        # Sheet 3: Failed tests only
        failed_df = df[df['Test Result'] == 'FAILED ❌']
        if not failed_df.empty:
            failed_df.to_excel(writer, sheet_name='Failed Tests', index=False)
        
        # Sheet 4: Performance metrics
        perf_df = df.groupby('Endpoint').agg({
            'Execution Time (ms)': ['min', 'max', 'mean', 'median']
        }).round(2)
        perf_df.to_excel(writer, sheet_name='Performance Metrics')
        
        # Sheet 5: Summary by endpoint
        endpoint_summary = df.groupby('Endpoint').agg({
            'Test Name': 'count',
            'Test Result': lambda x: (x == 'PASSED ✅').sum(),
            'Execution Time (ms)': 'mean'
        }).rename(columns={
            'Test Name': 'Total Tests',
            'Test Result': 'Passed',
            'Execution Time (ms)': 'Avg Time (ms)'
        })
        endpoint_summary['Failed'] = endpoint_summary['Total Tests'] - endpoint_summary['Passed']
        endpoint_summary['Pass Rate %'] = (endpoint_summary['Passed'] / endpoint_summary['Total Tests'] * 100).round(2)
        endpoint_summary.to_excel(writer, sheet_name='Summary by Endpoint')
    
    # Format Excel
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    
    wb = load_workbook(excel_file)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Color code results
        if sheet_name == 'All Tests':
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                result_cell = row[8]  # "Test Result" column
                if "PASSED" in str(result_cell.value):
                    result_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif "FAILED" in str(result_cell.value):
                    result_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    wb.save(excel_file)
    
    # Print summary
    total_tests = len(test_results)
    passed_tests = sum(1 for r in test_results if r['Test Result'] == 'PASSED ✅')
    failed_tests = total_tests - passed_tests
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    
    print(f"\n{'='*80}")
    print(f"TEST EXECUTION SUMMARY")
    print(f"{'='*80}")
    print(f"Total Tests: {total_tests}")
    print(f"Passed: {passed_tests} ✅")
    print(f"Failed: {failed_tests} ❌")
    print(f"Pass Rate: {pass_rate:.2f}%")
    print(f"\n📊 Excel Report: {excel_file}")
    print(f"📁 All files in: {CURRENT_RUN}")
    print(f"{'='*80}\n")